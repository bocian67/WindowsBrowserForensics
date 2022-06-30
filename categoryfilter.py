# -*- encoding: utf-8 -*-
from openpyxl import load_workbook


class CategoryFilter:
    def __init__(self, whitelist=None, blacklist=None, use_whitelist=True):
        if blacklist is None:
            blacklist = []
        if whitelist is None:
            whitelist = []
        self.excel_file = "C:/Users/wladi/Documents/Windows Browser Forensics/Linksammlung.xlsx"
        self.whitelist = whitelist
        self.blacklist = blacklist
        self.use_whitelist = use_whitelist
        self.workbook = load_workbook(filename=self.excel_file, read_only=True)

    def get_categories_from_source(self):
        categories = []
        for sheet in self.workbook.worksheets:
            categories.append(sheet.title)
        return categories

    def add_category_to_whitelist(self, category):
        ws = self.workbook[category]
        for row in range(2, ws.max_row+1):
            self.whitelist.append(ws["B" + str(row)].value)

    def add_category_to_blacklist(self, category):
        ws = self.workbook[category]
        for row in range(2, ws.max_row + 1):
            self.blacklist.append(ws["B" + str(row)].value)

    def get_filter_list(self):
        if self.use_whitelist:
            return self.whitelist
        else:
            return self.blacklist
